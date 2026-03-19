"""
Customer Churn Prediction - Flask Backend API
===============================================
Serves the dashboard and provides REST API endpoints for
real-time churn prediction, customer analytics, and batch processing.
"""

import os
import json
import io
import csv
import uuid
import secrets
import numpy as np
import pandas as pd
import joblib
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, make_response
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Load environment variables
load_dotenv()

# ── Admin Configuration ──────────────────────────────────────────────────
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "yugspam09@gmail.com")
active_sessions = {}  # In-memory session store
reset_tokens = {}     # In-memory store for reset tokens: {token: {"username": username, "expires": datetime}}

# ── Paths ────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model.pkl")
ENCODERS_PATH = os.path.join(BASE_DIR, "label_encoders.pkl")
SCALER_PATH = os.path.join(BASE_DIR, "scaler.pkl")
FEATURES_PATH = os.path.join(BASE_DIR, "feature_columns.json")
METRICS_PATH = os.path.join(BASE_DIR, "model_metrics.json")
STATS_PATH = os.path.join(BASE_DIR, "dataset_stats.json")
DATA_PATH = os.path.join(BASE_DIR, "customer_churn_500_dataset.csv")

# ── Flask App ────────────────────────────────────────────────────────────
app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))

# ── SMTP Configuration ───────────────────────────────────────────────────
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USERNAME = os.getenv("SMTP_USERNAME")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

# ── Authentication Decorator ─────────────────────────────────────────────
def require_auth(f):
    """Decorator to protect routes requiring authentication."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check session token from header or cookie
        auth_header = request.headers.get('Authorization')
        session_token = None
        
        if auth_header and auth_header.startswith('Bearer '):
            session_token = auth_header.split(' ')[1]
        elif 'session_token' in session:
            session_token = session['session_token']
        
        if not session_token or session_token not in active_sessions:
            # For API routes, return 401
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized", "redirect": "/login"}), 401
            # For page routes, redirect to login
            return redirect(url_for('admin_login'))
        
        return f(*args, **kwargs)
    return decorated_function

# ── Load ML Artifacts ────────────────────────────────────────────────────
model = joblib.load(MODEL_PATH)
label_encoders = joblib.load(ENCODERS_PATH)
scaler = joblib.load(SCALER_PATH)

with open(FEATURES_PATH, "r") as f:
    feature_columns = json.load(f)

with open(METRICS_PATH, "r") as f:
    model_metrics = json.load(f)

with open(STATS_PATH, "r") as f:
    dataset_stats = json.load(f)

# Load raw dataset for customer exploration
df_raw = pd.read_csv(DATA_PATH)

# ── Retention Strategies Database ────────────────────────────────────────
RETENTION_STRATEGIES = {
    "high_monthly_charges": {
        "title": "Offer Competitive Pricing",
        "description": "Customer's monthly charges are above average. Consider offering a loyalty discount of 15-25% or a bundled package deal.",
        "icon": "dollar-sign",
        "priority": "high"
    },
    "month_to_month_contract": {
        "title": "Incentivize Long-Term Contract",
        "description": "Customer is on a month-to-month contract. Offer a discounted annual plan with additional perks like free premium channels for 3 months.",
        "icon": "file-contract",
        "priority": "high"
    },
    "no_tech_support": {
        "title": "Provide Premium Tech Support",
        "description": "Customer lacks tech support. Offer complimentary tech support for 6 months to improve their experience and reduce friction.",
        "icon": "headset",
        "priority": "medium"
    },
    "no_online_security": {
        "title": "Add Security Package",
        "description": "Customer has no online security. Bundle a free security package to increase perceived value and stickiness.",
        "icon": "shield",
        "priority": "medium"
    },
    "fiber_optic_user": {
        "title": "Fiber Optic Retention Package",
        "description": "Fiber optic customers churn more frequently. Offer speed upgrades or a premium entertainment bundle at current pricing.",
        "icon": "wifi",
        "priority": "medium"
    },
    "low_tenure": {
        "title": "New Customer Onboarding Program",
        "description": "Customer is relatively new. Assign a dedicated account manager and provide a personalized welcome kit with usage tips.",
        "icon": "user-plus",
        "priority": "high"
    },
    "no_online_backup": {
        "title": "Free Cloud Backup Trial",
        "description": "Offer a 3-month free trial of online backup services to increase engagement and service dependency.",
        "icon": "cloud",
        "priority": "low"
    },
    "electronic_check_payment": {
        "title": "Autopay Incentive",
        "description": "Electronic check users churn more. Offer a $5/month discount for switching to automatic credit card or bank transfer payments.",
        "icon": "credit-card",
        "priority": "medium"
    },
    "no_device_protection": {
        "title": "Device Protection Plan",
        "description": "Offer a complimentary device protection plan for 6 months to increase service stickiness.",
        "icon": "mobile",
        "priority": "low"
    },
    "paperless_billing": {
        "title": "Billing Communication Enhancement",
        "description": "Ensure clear, detailed billing statements. Send monthly usage summaries highlighting value received vs. cost.",
        "icon": "envelope",
        "priority": "low"
    },
    "senior_citizen": {
        "title": "Senior-Friendly Support Program",
        "description": "Provide dedicated senior support line, simplified billing, and in-home setup assistance.",
        "icon": "heart",
        "priority": "medium"
    },
    "multiple_streaming": {
        "title": "Entertainment Loyalty Bonus",
        "description": "Customer uses streaming services. Offer exclusive content access or a partner streaming service subscription at no extra cost.",
        "icon": "tv",
        "priority": "low"
    }
}


def build_feature_vector(data: dict) -> pd.DataFrame:
    """
    Take raw customer data (from the prediction form) and build a
    feature vector matching the trained model's expected input.
    """
    # Map form fields to dataframe columns (matching new dataset structure)
    row = {
        "gender": data.get("gender", "Male"),
        "senior_citizen": 1 if data.get("senior_citizen", "No") == "Yes" else 0,
        "tenure_months": int(data.get("tenure_months", 1)),
        "contract_type": data.get("contract", "Month-to-month"),
        "internet_service": data.get("internet_service", "No"),
        "monthly_charges": float(data.get("monthly_charges", 50)),
        "total_charges": float(data.get("total_charges", 50)),
        "payment_method": data.get("payment_method", "Electronic check"),
    }

    df = pd.DataFrame([row])

    # ── Feature engineering (must match train_model.py exactly) ──
    # Tenure buckets
    df["tenure_bucket"] = pd.cut(
        df["tenure_months"],
        bins=[0, 12, 24, 48, 72, 200],
        labels=["0-12", "13-24", "25-48", "49-72", "72+"],
    )

    # Avg monthly charge relative to tenure
    df["avg_charge_per_month"] = np.where(
        df["tenure_months"] > 0,
        df["total_charges"] / df["tenure_months"],
        df["monthly_charges"],
    )

    # Has internet
    if "internet_service" in df.columns:
        df["has_internet"] = df["internet_service"].apply(lambda x: 0 if x == "No" else 1)

    # Encode categoricals
    cat_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()
    for col in cat_cols:
        if col in label_encoders:
            le = label_encoders[col]
            # Handle unseen labels gracefully
            known = set(le.classes_)
            df[col] = df[col].apply(lambda x: x if x in known else le.classes_[0])
            df[col] = le.transform(df[col].astype(str))
        else:
            df[col] = 0

    # Ensure columns match model's expectations
    for col in feature_columns:
        if col not in df.columns:
            df[col] = 0

    df = df[feature_columns]

    # Scale
    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    df[num_cols] = scaler.transform(df[num_cols])

    df = df.fillna(0)

    return df


def get_retention_strategies(data: dict, churn_prob: float) -> list:
    """Determine personalized retention strategies based on customer profile."""
    strategies = []

    monthly_charges = float(data.get("monthly_charges", 0))
    avg_monthly = dataset_stats.get("avg_monthly_charges", 65)

    if monthly_charges > avg_monthly * 1.1:
        strategies.append(RETENTION_STRATEGIES["high_monthly_charges"])

    if data.get("contract", "") == "Month-to-month":
        strategies.append(RETENTION_STRATEGIES["month_to_month_contract"])

    tenure = int(data.get("tenure_months", 0))
    if tenure < 12:
        strategies.append(RETENTION_STRATEGIES["low_tenure"])

    if data.get("internet_service", "") == "Fiber optic":
        strategies.append(RETENTION_STRATEGIES["fiber_optic_user"])

    if data.get("payment_method", "") == "Electronic check":
        strategies.append(RETENTION_STRATEGIES["electronic_check_payment"])

    if data.get("senior_citizen", "No") == "Yes":
        strategies.append(RETENTION_STRATEGIES["senior_citizen"])

    # Sort by priority
    priority_order = {"high": 0, "medium": 1, "low": 2}
    strategies.sort(key=lambda s: priority_order.get(s["priority"], 3))

    return strategies[:5]  # Top 5 strategies


def get_risk_level(prob: float) -> dict:
    """Classify churn probability into risk levels."""
    if prob >= 0.75:
        return {"level": "Critical", "color": "#ff2d55", "action": "Immediate intervention required"}
    elif prob >= 0.50:
        return {"level": "High", "color": "#ff9500", "action": "Proactive retention outreach needed"}
    elif prob >= 0.25:
        return {"level": "Medium", "color": "#ffcc00", "action": "Monitor closely and engage periodically"}
    else:
        return {"level": "Low", "color": "#34c759", "action": "Standard engagement is sufficient"}


def get_contributing_factors(data: dict) -> list:
    """Identify top contributing factors for the churn prediction."""
    factors = []

    if data.get("contract", "") == "Month-to-month":
        factors.append({"factor": "Month-to-month contract", "impact": "high", "description": "Short-term contracts have 42% churn rate"})

    tenure = int(data.get("tenure_months", 0))
    if tenure < 12:
        factors.append({"factor": f"Low tenure ({tenure} months)", "impact": "high", "description": "New customers (<12 months) are 3x more likely to churn"})

    monthly = float(data.get("monthly_charges", 0))
    if monthly > 70:
        factors.append({"factor": f"High monthly charges (${monthly:.2f})", "impact": "medium", "description": "Above-average charges correlate with higher churn"})

    if data.get("internet_service", "") == "Fiber optic":
        factors.append({"factor": "Fiber optic service", "impact": "medium", "description": "Fiber optic users show higher churn rates"})

    if data.get("payment_method", "") == "Electronic check":
        factors.append({"factor": "Electronic check payment", "impact": "medium", "description": "Electronic check users have higher churn rates"})

    if data.get("senior_citizen", "No") == "Yes":
        factors.append({"factor": "Senior citizen", "impact": "medium", "description": "Senior citizens have higher churn rates"})

    impact_order = {"high": 0, "medium": 1, "low": 2}
    factors.sort(key=lambda f: impact_order.get(f["impact"], 3))

    return factors[:5]


# ══════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════

@app.route("/login")
def admin_login():
    """Serve the admin login page."""
    # If already logged in, redirect to dashboard
    session_token = session.get('session_token')
    if session_token and session_token in active_sessions:
        return redirect(url_for('index'))
    return render_template("admin.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    """Handle admin login authentication."""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            return jsonify({"success": False, "message": "Username and password are required"}), 400
        
        # Validate credentials
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            # Generate session token
            token = secrets.token_urlsafe(32)
            active_sessions[token] = {
                "username": username,
                "created_at": pd.Timestamp.now().isoformat()
            }
            
            # Store in session cookie
            session['session_token'] = token
            
            return jsonify({
                "success": True,
                "message": "Login successful",
                "token": token
            })
        else:
            return jsonify({"success": False, "message": "Invalid credentials"}), 401
            
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/forgot-password", methods=["POST"])
def forgot_password():
    """Handle password reset request."""
    try:
        data = request.get_json()
        print(f"DEBUG: Forgot password request received: {data}")
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        
        if not username or not email:
            return jsonify({"success": False, "message": "Username and email are required"}), 400
            
        if username == ADMIN_USERNAME and email == ADMIN_EMAIL:
            # Generate reset token
            token = secrets.token_urlsafe(32)
            expires = datetime.now() + timedelta(hours=1)
            reset_tokens[token] = {
                "username": username,
                "expires": expires
            }
            
            # Send email
            reset_link = f"{request.url_root.rstrip('/')}/reset-password/{token}"
            if send_reset_email(email, reset_link):
                return jsonify({"success": True, "message": "Password reset link sent to your email"})
            else:
                return jsonify({"success": False, "message": "Failed to send email. Check SMTP configuration."}), 500
        else:
            return jsonify({"success": False, "message": "Invalid username or email"}), 401
            
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


def send_reset_email(receiver_email, reset_link):
    """Send a real reset email using SMTP."""
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        print("ERROR: SMTP credentials not found in environment variables.")
        return False
        
    try:
        message = MIMEMultipart("alternative")
        message["Subject"] = "ChurnGuard - Password Reset Request"
        message["From"] = f"ChurnGuard <{SMTP_USERNAME}>"
        message["To"] = receiver_email

        text = f"Hello,\n\nYou requested a password reset for ChurnGuard.\nClick the link below to reset your password:\n{reset_link}\n\nThis link will expire in 1 hour."
        html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
                <h2 style="color: #6366f1;">ChurnGuard</h2>
                <p>Hello,</p>
                <p>You requested a password reset for your admin account. Click the button below to set a new password:</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{reset_link}" style="background-color: #6366f1; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold;">Reset Password</a>
                </div>
                <p>If the button doesn't work, copy and paste this link into your browser:</p>
                <p style="word-break: break-all; color: #8b5cf6;">{reset_link}</p>
                <p>This link will expire in 1 hour. If you didn't request this, you can safely ignore this email.</p>
                <hr style="border: 0; border-top: 1px solid #eee; margin: 20px 0;">
                <p style="font-size: 0.8rem; color: #888;">&copy; 2026 ChurnGuard. All rights reserved.</p>
            </div>
        </body>
        </html>
        """
        
        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")
        message.attach(part1)
        message.attach(part2)

        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_USERNAME, receiver_email, message.as_string())
        
        return True
    except Exception as e:
        print(f"SMTP Error: {e}")
        return False


@app.route("/reset-password/<token>")
def reset_password_page(token):
    """Serve the reset password page if token is valid."""
    if token in reset_tokens:
        token_data = reset_tokens[token]
        if datetime.now() < token_data["expires"]:
            return render_template("reset_password.html", token=token)
        else:
            del reset_tokens[token]
    
    return render_template("admin.html", error="Password reset link has expired or is invalid")


@app.route("/api/reset-password", methods=["POST"])
def api_reset_password():
    """Handle the actual password reset."""
    try:
        data = request.get_json()
        token = data.get('token')
        new_password = data.get('password')
        
        if not token or not new_password:
            return jsonify({"success": False, "message": "Token and new password are required"}), 400
            
        if token in reset_tokens:
            token_data = reset_tokens[token]
            if datetime.now() < token_data["expires"]:
                global ADMIN_PASSWORD
                ADMIN_PASSWORD = new_password
                del reset_tokens[token]
                return jsonify({"success": True, "message": "Password updated successfully. You can now login with your new password."})
            else:
                del reset_tokens[token]
                return jsonify({"success": False, "message": "Reset link has expired"}), 400
        else:
            return jsonify({"success": False, "message": "Invalid reset token"}), 400
            
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/api/logout", methods=["POST"])
def api_logout():
    """Handle admin logout."""
    session_token = session.get('session_token')
    if session_token and session_token in active_sessions:
        del active_sessions[session_token]
    session.pop('session_token', None)
    return jsonify({"success": True, "message": "Logged out successfully"})


@app.route("/api/verify-session", methods=["POST"])
def verify_session():
    """Verify if a session token is valid."""
    try:
        data = request.get_json()
        token = data.get('token') or session.get('session_token')
        
        if token and token in active_sessions:
            return jsonify({"valid": True, "username": active_sessions[token]["username"]})
        else:
            return jsonify({"valid": False}), 401
    except Exception as e:
        return jsonify({"valid": False, "error": str(e)}), 401


@app.route("/")
@require_auth
def index():
    """Serve the main dashboard."""
    return render_template("index.html")


@app.route("/api/predict", methods=["POST"])
@require_auth
def predict():
    """Predict churn for a single customer."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        feature_vector = build_feature_vector(data)
        proba = model.predict_proba(feature_vector)[0]
        churn_prob = float(proba[1])

        risk = get_risk_level(churn_prob)
        strategies = get_retention_strategies(data, churn_prob)
        factors = get_contributing_factors(data)

        result = {
            "churn_probability": round(churn_prob * 100, 1),
            "risk_level": risk["level"],
            "risk_color": risk["color"],
            "risk_action": risk["action"],
            "contributing_factors": factors,
            "retention_strategies": strategies,
            "prediction": "Will Churn" if churn_prob >= 0.5 else "Will Stay",
        }

        return jsonify(result)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/stats")
@require_auth
def stats():
    """Return dataset statistics and model metrics."""
    return jsonify({
        "dataset": dataset_stats,
        "model": model_metrics,
    })


@app.route("/api/customers")
@require_auth
def customers():
    """Return paginated customer list with churn info."""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    search = request.args.get("search", "", type=str)
    sort_by = request.args.get("sort_by", "churn", type=str)
    sort_dir = request.args.get("sort_dir", "desc", type=str)

    df = df_raw.copy()

    # Search
    if search:
        mask = df.apply(lambda row: search.lower() in str(row).lower(), axis=1)
        df = df[mask]

    # Map frontend sort values to DataFrame column names
    sort_mapping = {
        "Churn Score": "churn",
        "Monthly Charges": "monthly_charges",
        "Tenure Months": "tenure_months",
        "Total Charges": "total_charges"
    }
    
    # Use mapped column name if available, otherwise use original
    sort_column = sort_mapping.get(sort_by, sort_by)
    
    # Sort
    if sort_column in df.columns:
        df = df.sort_values(sort_column, ascending=(sort_dir == "asc"))

    total = len(df)
    total_pages = max(1, (total + per_page - 1) // per_page)
    start = (page - 1) * per_page
    end = start + per_page

    page_df = df.iloc[start:end]

    customers_list = []
    for _, row in page_df.iterrows():
        # Convert senior_citizen from 0/1 to No/Yes for display
        senior_display = "Yes" if row.get("senior_citizen", 0) == 1 else "No"
        # Convert churn from 0/1 to No/Yes for display
        churn_label = "Yes" if row.get("churn", 0) == 1 else "No"
        customers_list.append({
            "customer_id": row.get("customer_id", ""),
            "gender": row.get("gender", ""),
            "senior_citizen": senior_display,
            "tenure_months": int(row.get("tenure_months", 0)),
            "contract": row.get("contract_type", ""),
            "monthly_charges": float(row.get("monthly_charges", 0)),
            "total_charges": str(row.get("total_charges", "")),
            "churn_label": churn_label,
            "churn_score": int(row.get("churn", 0) * 100),  # Use churn as score (0 or 100)
            "internet_service": row.get("internet_service", ""),
            "payment_method": row.get("payment_method", ""),
        })

    return jsonify({
        "customers": customers_list,
        "page": page,
        "per_page": per_page,
        "total": total,
        "total_pages": total_pages,
    })


@app.route("/api/customer/<customer_id>")
@require_auth
def customer_detail(customer_id):
    """Return detailed customer profile with prediction."""
    row = df_raw[df_raw["customer_id"] == customer_id]
    if row.empty:
        return jsonify({"error": "Customer not found"}), 404

    row = row.iloc[0]
    # Convert senior_citizen from 0/1 to Yes/No for form processing
    senior_val = "Yes" if row.get("senior_citizen", 0) == 1 else "No"
    data = {
        "gender": row.get("gender", "Male"),
        "senior_citizen": senior_val,
        "tenure_months": str(int(row.get("tenure_months", 1))),
        "contract": row.get("contract_type", "Month-to-month"),
        "internet_service": row.get("internet_service", "No"),
        "payment_method": row.get("payment_method", "Electronic check"),
        "monthly_charges": str(float(row.get("monthly_charges", 50))),
        "total_charges": str(row.get("total_charges", "50")),
    }

    # Run prediction
    feature_vector = build_feature_vector(data)
    proba = model.predict_proba(feature_vector)[0]
    churn_prob = float(proba[1])

    risk = get_risk_level(churn_prob)
    strategies = get_retention_strategies(data, churn_prob)
    factors = get_contributing_factors(data)

    # Convert churn from 0/1 to Yes/No for display
    churn_label = "Yes" if row.get("churn", 0) == 1 else "No"

    return jsonify({
        "customer": {
            "customer_id": row.get("customer_id", ""),
            "churn_label": churn_label,
            **data,
        },
        "prediction": {
            "churn_probability": round(churn_prob * 100, 1),
            "risk_level": risk["level"],
            "risk_color": risk["color"],
            "risk_action": risk["action"],
            "contributing_factors": factors,
            "retention_strategies": strategies,
        }
    })


@app.route("/api/batch-predict", methods=["POST"])
@require_auth
def batch_predict():
    """Accept CSV upload and return batch predictions."""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        if not file.filename.endswith(".csv"):
            return jsonify({"error": "File must be a CSV"}), 400

        df_upload = pd.read_csv(file)
        results = []

        for _, row in df_upload.iterrows():
            # Convert senior_citizen from 0/1 to Yes/No string
            senior_val = "Yes" if row.get("senior_citizen", 0) == 1 else "No"
            data = {
                "gender": row.get("gender", "Male"),
                "senior_citizen": senior_val,
                "tenure_months": str(int(row.get("tenure_months", 1))),
                "contract": row.get("contract_type", "Month-to-month"),
                "internet_service": row.get("internet_service", "No"),
                "payment_method": row.get("payment_method", "Electronic check"),
                "monthly_charges": str(float(row.get("monthly_charges", 50))),
                "total_charges": str(row.get("total_charges", "50")),
            }

            try:
                feature_vector = build_feature_vector(data)
                proba = model.predict_proba(feature_vector)[0]
                churn_prob = float(proba[1])
                risk = get_risk_level(churn_prob)

                results.append({
                    "customer_id": row.get("customer_id", "N/A"),
                    "churn_probability": round(churn_prob * 100, 1),
                    "risk_level": risk["level"],
                    "prediction": "Will Churn" if churn_prob >= 0.5 else "Will Stay",
                })
            except Exception:
                results.append({
                    "customer_id": row.get("customer_id", "N/A"),
                    "churn_probability": None,
                    "risk_level": "Error",
                    "prediction": "Error processing",
                })

        # Summary
        valid = [r for r in results if r["churn_probability"] is not None]
        summary = {
            "total_processed": len(results),
            "predicted_churn": len([r for r in valid if r["churn_probability"] >= 50]),
            "predicted_stay": len([r for r in valid if r["churn_probability"] < 50]),
            "avg_churn_probability": round(np.mean([r["churn_probability"] for r in valid]), 1) if valid else 0,
            "errors": len(results) - len(valid),
        }

        return jsonify({
            "results": results,
            "summary": summary,
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Report Generation Helpers ─────────────────────────────────────────────

def get_report_data(period="weekly"):
    """
    Get a subset of data for the report.
    Since we don't have real dates, we'll simulate this by picking 
    high-risk customers based on specific criteria.
    """
    # Create copy of dataframe
    df_sorted = df_raw.copy()
    
    if period == "weekly":
        # Get 50 highest risk (simulating most recent high risk)
        data = df_sorted.sort_values("tenure_months", ascending=True).head(50)
    else:
        # Get a larger sample to represent the month
        data = df_sorted.sample(min(100, len(df_sorted)))
    
    return data

def create_excel_report(df, period):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{period.capitalize()} Churn Report"

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Main Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"CHURNGUARD - {period.upper()} CHURN ANALYSIS"
    title_cell.font = Font(size=16, bold=True, color="374151")
    title_cell.alignment = center_align

    # Report Stats Subtitle
    ws.merge_cells('A2:I2')
    sub_cell = ws['A2']
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Analysis Segment: {len(df)} Records"
    sub_cell.font = Font(size=11, color="6B7280")
    sub_cell.alignment = center_align

    ws.append([]) # Empty row

    # Headers
    headers = ["Customer ID", "Gender", "Tenure", "Contract", "Monthly Charges", "Total Charges", "Internet", "Payment", "Churn?"]
    ws.append(headers)
    
    header_row = ws[ws.max_row]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    for _, row in df.iterrows():
        ws.append([
            row['customer_id'], row['gender'], f"{row['tenure_months']} mo",
            row['contract_type'], f"${row['monthly_charges']:.2f}",
            f"${row['total_charges']:.2f}", row['internet_service'],
            row['payment_method'], "CHURNED" if int(row['churn']) == 1 else "ACTIVE"
        ])
        
        current_row = ws[ws.max_row]
        for cell in current_row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter if hasattr(col[0], 'column_letter') else get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 3

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

class PDFReport(FPDF):
    def header(self):
        # Brand Header
        self.set_fill_color(99, 102, 241) # Indigo
        self.rect(0, 0, 210, 40, 'F')
        self.set_xy(10, 12)
        self.set_font('helvetica', 'B', 24)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, 'ChurnGuard', ln=True)
        self.set_font('helvetica', '', 10)
        self.cell(0, 6, 'Executive Business Intelligence Report', ln=True)
        self.ln(12)

    def footer(self):
        self.set_y(-20)
        self.set_font('helvetica', 'I', 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f'Page {self.page_no()} | Generated {datetime.now().strftime("%Y-%m-%d")}', 0, 0, 'L')

def create_pdf_report(df, period):
    pdf = PDFReport()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    pdf.set_font('helvetica', 'B', 16)
    pdf.set_text_color(31, 41, 55)
    pdf.cell(0, 15, f"{period.capitalize()} Churn Segment Analysis", ln=True)
    
    # Table Results
    pdf.set_fill_color(99, 102, 241)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(40, 10, "Customer ID", 1, 0, 'C', True)
    pdf.cell(25, 10, "Tenure", 1, 0, 'C', True)
    pdf.cell(50, 10, "Contract", 1, 0, 'C', True)
    pdf.cell(45, 10, "Monthly Charges", 1, 0, 'C', True)
    pdf.cell(30, 10, "Status", 1, 1, 'C', True)

    pdf.set_text_color(31, 41, 55)
    pdf.set_font('helvetica', '', 8)
    
    count = 0
    for _, row in df.iterrows():
        if count >= 30: break
        bg = (245, 247, 250) if count % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*bg)
        pdf.cell(40, 8, str(row['customer_id']), 1, 0, 'C', True)
        pdf.cell(25, 8, f"{row['tenure_months']} mo", 1, 0, 'C', True)
        pdf.cell(50, 8, str(row['contract_type']), 1, 0, 'L', True)
        pdf.cell(45, 8, f"${row['monthly_charges']:.2f}", 1, 0, 'R', True)
        pdf.cell(30, 8, "CHURNED" if int(row['churn']) == 1 else "ACTIVE", 1, 1, 'C', True)
        count += 1

    return pdf.output(dest='S')

# ── Reporting API ────────────────────────────────────────────────────────

@app.route('/api/reports/download', methods=['GET'])
@require_auth
def download_churn_report():
    """Download churn reports in Excel or PDF format."""
    fmt = request.args.get('format', 'pdf').lower()
    period = request.args.get('period', 'weekly').lower()
    
    if period not in ['weekly', 'monthly']:
        return jsonify({"error": "Invalid analysis period"}), 400
        
    df_subset = get_report_data(period)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    try:
        if fmt == 'excel':
            output = create_excel_report(df_subset, period)
            filename = f"ChurnGuard_{period.capitalize()}_Report_{timestamp}.xlsx"
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            pdf_data = create_pdf_report(df_subset, period)
            if isinstance(pdf_data, str):
                pdf_data = pdf_data.encode('latin-1')
            
            filename = f"ChurnGuard_{period.capitalize()}_Report_{timestamp}.pdf"
            return send_file(
                io.BytesIO(pdf_data),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            )
    except Exception as e:
        print(f"Report error: {str(e)}")
        return jsonify({"error": "Failed to generate report"}), 500


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("  Customer Churn Prediction - Server")
    print("=" * 60)
    print(f"  Model: {model_metrics.get('best_model', 'Unknown')}")
    print(f"  Features: {len(feature_columns)}")
    print(f"  Dataset: {dataset_stats.get('total_customers', 0)} customers")
    print(f"  Churn Rate: {dataset_stats.get('churn_rate', 0)}%")
    print("=" * 60 + "\n")

    app.run(debug=True, host="127.0.0.1", port=5000)
